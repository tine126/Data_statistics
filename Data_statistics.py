import os
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

def extract_report_data(folder_path):
    # 尝试读取 report.json 文件（用于获取 plate 和 timestamp）
    report_file = os.path.join(folder_path, 'report.json')
    plate, timestamp = None, None
    
    if os.path.exists(report_file):
        try:
            with open(report_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                plate = data.get("plate", None)
                timestamp = data.get("timestamp", None)
        except UnicodeDecodeError:
            print(f"Error decoding {report_file}, skipping...")

    # 尝试读取 reports.json 文件（用于获取 length, width, height 和 value）
    reports_file = os.path.join(folder_path, 'reports.json')
    length, width, height, value = None, None, None, None
    
    if os.path.exists(reports_file):
        try:
            with open(reports_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                length = data["dimensions"].get("length", None)
                width = data["dimensions"].get("width", None)
                height = data["dimensions"].get("height", None)
                value = data["volume"].get("value", None)
        except UnicodeDecodeError:
            print(f"Error decoding {reports_file}, skipping...")

    return length, width, height, value, plate, timestamp

def create_excel_report(output_file):
    # 创建 Excel 工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Data"

    # 写入表头
    header = ["Folder Name", "Plate", "Timestamp", "Length", "Width", "Height", "Value", "Folder Link"]
    ws.append(header)

    # 设置列宽
    ws.column_dimensions['A'].width = 20  # Folder Name
    ws.column_dimensions['B'].width = 15  # Plate
    ws.column_dimensions['C'].width = 30  # Timestamp
    ws.column_dimensions['D'].width = 15  # Length
    ws.column_dimensions['E'].width = 15  # Width
    ws.column_dimensions['F'].width = 15  # Height
    ws.column_dimensions['G'].width = 15  # Value
    ws.column_dimensions['H'].width = 50  # Folder Link

    # 设置居中对齐和边框
    alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # 应用格式到表头
    for cell in ws[1]:
        cell.alignment = alignment
        cell.border = thin_border

    # 获取当前文件夹下的所有子文件夹
    base_dir = os.getcwd()
    for folder_name in os.listdir(base_dir):
        folder_path = os.path.join(base_dir, folder_name)

        if os.path.isdir(folder_path):
            # 获取 report.json 和 reports.json 中的数据
            length, width, height, value, plate, timestamp = extract_report_data(folder_path)

            if length is not None and width is not None and height is not None and value is not None:
                # 添加行数据，包括 plate 和 timestamp
                row = [folder_name, plate, timestamp, length, width, height, value]
                ws.append(row)

                # 为文件夹添加超链接
                folder_link = f"file:///{folder_path.replace(os.sep, '/')}"
                cell = ws.cell(row=ws.max_row, column=8)
                cell.value = folder_link
                cell.hyperlink = folder_link

                # 对所有单元格应用居中对齐和边框
                for cell in ws[ws.max_row]:
                    cell.alignment = alignment
                    cell.border = thin_border

    # 保存 Excel 文件
    wb.save(output_file)
    print(f"Report saved as {output_file}")

if __name__ == "__main__":
    output_excel_file = "Data_statistics.xlsx"
    create_excel_report(output_excel_file)
